import os
import threading
import tkinter as tk
from datetime import datetime
from tkinter import messagebox

from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.service import Service
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.microsoft import EdgeChromiumDriverManager

EXCEL_FILE = r"C:\Users\Vinicius\Documents\PY\captador_temperatura\temperatura.xlsx"
URL_TEMPO = "https://www.tempo.com/sao-paulo.htm"
XPATH_TEMPERATURA = "//*[@id='d_hub_1']/div[1]/div/div/div/div/span[1]"
XPATH_UMIDADE = (
    "/html/body/main/div[1]/div/section[2]/div/ul/li[1]/span/span[3]/span[2]/span[1]"
)


# Inicializa o WebDriver do Edge automaticamente
def iniciar_driver():
    try:
        return webdriver.Edge(service=Service(EdgeChromiumDriverManager().install()))
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao iniciar WebDriver:\n{e}")
        return None


# Obtém a temperatura e umidade do ar de São Paulo
def obter_dados_climaticos():
    driver = iniciar_driver()
    if not driver:
        return None, None, None

    driver.get(URL_TEMPO)
    try:
        elemento_temp = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.XPATH, XPATH_TEMPERATURA))
        )
        elemento_umi = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.XPATH, XPATH_UMIDADE))
        )
        temperatura = "".join(filter(str.isdigit, elemento_temp.text))
        umidade = "".join(filter(str.isdigit, elemento_umi.text))
        horario = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        return horario, temperatura, umidade
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao obter dados climáticos:\n{e}")
        return None, None, None
    finally:
        driver.quit()


# Salva os dados da temperatura e umidade em um arquivo Excel
def salvar_excel(dados):
    # Cria o diretório, caso não exista
    os.makedirs(os.path.dirname(EXCEL_FILE), exist_ok=True)

    wb = load_workbook(EXCEL_FILE) if os.path.exists(EXCEL_FILE) else Workbook()
    ws = wb.active

    if not os.path.exists(EXCEL_FILE):
        ws.append(["Horário", "Temperatura (°C)", "Umidade (%)"])

    ws.append(dados)
    wb.save(EXCEL_FILE)


# Inicia a busca dos dados em uma nova thread
def buscar_dados_climaticos():
    threading.Thread(target=registrar_dados_climaticos, daemon=True).start()


# Captura os dados climáticos, salva no Excel e atualiza a interface gráfica
def registrar_dados_climaticos():
    horario, temperatura, umidade = obter_dados_climaticos()
    if horario:
        salvar_excel((horario, temperatura, umidade))
        resultado_label.config(text=f"São Paulo: {temperatura}°C | Umidade: {umidade}%")
        messagebox.showinfo(
            "Sucesso",
            f"Dados registrados:\n{horario}\nTemperatura: {temperatura}°C\nUmidade: {umidade}%",
        )


# Configuração da interface gráfica
janela = tk.Tk()
janela.title("Clima em São Paulo")
janela.geometry("350x180")
janela.resizable(False, False)

tk.Label(janela, text="Clima de São Paulo", font=("Arial", 14, "bold")).pack(pady=5)
tk.Button(
    janela, text="Buscar Dados", font=("Arial", 12), command=buscar_dados_climaticos
).pack(pady=10)
resultado_label = tk.Label(janela, text="", font=("Arial", 14))
resultado_label.pack(pady=10)

janela.mainloop()
